from django.http import HttpResponse
from drf_spectacular.types import OpenApiTypes
from drf_spectacular.utils import extend_schema
from openpyxl import Workbook
from rest_framework import viewsets
from rest_framework.decorators import api_view, permission_classes

from apps.bonuses.models import BonusAccount
from apps.orders.models import Order
from apps.products.models import Product
from apps.sales.models import Sale
from common.permissions import IsOwnerAdminManager
from common.views import CreatedByModelViewSet

from .models import ReportExport
from .serializers import ReportExportSerializer


def workbook_response(workbook, filename):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


def apply_common_filters(queryset, request):
    date_from = request.query_params.get("date_from")
    date_to = request.query_params.get("date_to")
    status = request.query_params.get("status")
    warehouse = request.query_params.get("warehouse")
    if date_from:
        queryset = queryset.filter(created_at__date__gte=date_from)
    if date_to:
        queryset = queryset.filter(created_at__date__lte=date_to)
    if status:
        queryset = queryset.filter(status=status)
    if warehouse:
        queryset = queryset.filter(warehouse_id=warehouse)
    return queryset


class ReportExportViewSet(CreatedByModelViewSet):
    queryset = ReportExport.objects.all().order_by("-created_at")
    serializer_class = ReportExportSerializer
    permission_classes = [IsOwnerAdminManager]


@extend_schema(tags=["Reports"], description="Download sales report as XLSX.", responses={(200, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"): OpenApiTypes.BINARY})
@api_view(["GET"])
@permission_classes([IsOwnerAdminManager])
def sales_report(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales"
    sheet.append(["ID", "Cashier", "Client", "Warehouse", "Status", "Subtotal", "Cost", "Discount", "Bonus spent", "Total", "Profit", "Created"])
    sales = apply_common_filters(Sale.objects.select_related("cashier", "client", "warehouse"), request)
    cashier = request.query_params.get("cashier")
    if cashier:
        sales = sales.filter(cashier_id=cashier)
    for sale in sales.order_by("-created_at"):
        sheet.append([
            sale.id,
            sale.cashier.username,
            sale.client.username if sale.client else "",
            sale.warehouse.name if sale.warehouse else "",
            sale.status,
            sale.subtotal,
            sale.cost_total,
            sale.discount_total,
            sale.bonus_spent,
            sale.total,
            sale.profit,
            sale.created_at.isoformat(),
        ])
    return workbook_response(workbook, "sales_report.xlsx")


@extend_schema(tags=["Reports"], description="Download inventory report as XLSX.", responses={(200, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"): OpenApiTypes.BINARY})
@api_view(["GET"])
@permission_classes([IsOwnerAdminManager])
def inventory_report(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Inventory"
    sheet.append(["ID", "SKU", "Name", "Brand", "Category", "Total stock", "Low threshold", "Retail price", "Cost price"])
    for product in Product.objects.select_related("brand", "category").order_by("name"):
        sheet.append([
            product.id,
            product.sku,
            product.name,
            product.brand.name,
            product.category.name,
            product.stock_quantity,
            product.low_stock_threshold,
            product.price,
            product.cost_price,
        ])
    return workbook_response(workbook, "inventory_report.xlsx")


@extend_schema(tags=["Reports"], description="Download orders report as XLSX.", responses={(200, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"): OpenApiTypes.BINARY})
@api_view(["GET"])
@permission_classes([IsOwnerAdminManager])
def orders_report(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Orders"
    sheet.append(["ID", "Client", "Warehouse", "Status", "Subtotal", "Delivery", "Discount", "Total", "Created"])
    orders = apply_common_filters(Order.objects.select_related("client", "warehouse"), request)
    client = request.query_params.get("client")
    if client:
        orders = orders.filter(client_id=client)
    for order in orders.order_by("-created_at"):
        sheet.append([
            order.id,
            order.client.username,
            order.warehouse.name if order.warehouse else "",
            order.status,
            order.subtotal,
            order.delivery_price,
            order.discount_total,
            order.total,
            order.created_at.isoformat(),
        ])
    return workbook_response(workbook, "orders_report.xlsx")


@extend_schema(tags=["Reports"], description="Download bonuses report as XLSX.", responses={(200, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"): OpenApiTypes.BINARY})
@api_view(["GET"])
@permission_classes([IsOwnerAdminManager])
def bonuses_report(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Bonuses"
    sheet.append(["Client", "Balance", "Updated"])
    for account in BonusAccount.objects.select_related("client").order_by("client__username"):
        sheet.append([account.client.username, account.balance, account.updated_at.isoformat()])
    return workbook_response(workbook, "bonuses_report.xlsx")
